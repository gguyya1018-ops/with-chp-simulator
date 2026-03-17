import webview
import json
import os
import sys
import glob
import shutil
import calendar
import ctypes
from openpyxl import load_workbook, Workbook



# PyInstaller exe: 리소스(html,css,js)는 번들 내부, 데이터는 exe 옆 폴더
if getattr(sys, 'frozen', False):
    RESOURCE_DIR = sys._MEIPASS          # 번들된 리소스 경로
    BASE_DIR = os.path.dirname(sys.executable)  # exe가 있는 폴더
else:
    RESOURCE_DIR = os.path.dirname(os.path.abspath(__file__))
    BASE_DIR = RESOURCE_DIR


class Api:
    # ─── 프로젝트 관리 ───
    def list_projects(self):
        data_dir = os.path.join(BASE_DIR, 'data')
        if not os.path.isdir(data_dir):
            return []
        return sorted([d for d in os.listdir(data_dir)
                       if os.path.isdir(os.path.join(data_dir, d))])

    def create_project(self, name):
        proj_dir = os.path.join(BASE_DIR, 'data', name)
        if os.path.exists(proj_dir):
            return {'error': '이미 존재하는 프로젝트입니다'}
        os.makedirs(proj_dir, exist_ok=True)
        return {'ok': True, 'path': proj_dir}

    def delete_project(self, name):
        proj_dir = os.path.join(BASE_DIR, 'data', name)
        if os.path.isdir(proj_dir):
            shutil.rmtree(proj_dir)
            return {'ok': True}
        return {'error': '폴더 없음'}

    # ─── CSV I/O ───
    def load_project(self, project_name):
        proj_dir = os.path.join(BASE_DIR, 'data', project_name)
        if not os.path.isdir(proj_dir):
            return {'error': f'프로젝트 폴더 없음: {proj_dir}'}
        result = {}
        for csv_path in glob.glob(os.path.join(proj_dir, '*.csv')):
            key = os.path.splitext(os.path.basename(csv_path))[0]
            try:
                with open(csv_path, 'r', encoding='utf-8-sig') as f:
                    result[key] = f.read()
            except Exception as e:
                result[key] = f'ERROR:{e}'
        return result

    def save_csv(self, project_name, file_key, content):
        proj_dir = os.path.join(BASE_DIR, 'data', project_name)
        os.makedirs(proj_dir, exist_ok=True)
        path = os.path.join(proj_dir, file_key + '.csv')
        with open(path, 'w', encoding='utf-8-sig', newline='') as f:
            f.write(content)
        return {'ok': True}

    def delete_csv(self, project_name, file_key):
        path = os.path.join(BASE_DIR, 'data', project_name, file_key + '.csv')
        if os.path.isfile(path):
            os.remove(path)
            return {'ok': True}
        return {'error': '파일 없음'}

    # ─── 파일 다이얼로그 ───
    def open_file_dialog(self, file_types):
        result = window.create_file_dialog(
            webview.OPEN_DIALOG, file_types=file_types)
        if result and len(result) > 0:
            path = result[0]
            if path.lower().endswith(('.xlsx', '.xls')):
                return self._read_excel(path)
            try:
                with open(path, 'r', encoding='utf-8-sig') as f:
                    return {'path': path, 'content': f.read()}
            except Exception as e:
                return {'error': str(e)}
        return None

    def _read_excel(self, path):
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            sheets = {}
            for name in wb.sheetnames:
                ws = wb[name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = ['' if c is None else str(c) for c in row]
                    rows.append(','.join(cells))
                if rows:
                    sheets[name] = '\n'.join(rows)
            wb.close()
            return {'path': path, 'sheets': sheets}
        except Exception as e:
            return {'error': str(e)}

    def export_excel(self, project_name, data_dict):
        """데이터를 Excel 파일로 내보내기"""
        result = window.create_file_dialog(
            webview.SAVE_DIALOG,
            save_filename=f'{project_name}_data.xlsx',
            file_types=('Excel Files (*.xlsx)',))
        if not result:
            return None
        path = result if isinstance(result, str) else result[0]
        if not path.lower().endswith('.xlsx'):
            path += '.xlsx'

        wb = Workbook()
        wb.remove(wb.active)

        for key, csv_text in data_dict.items():
            title = key[:31]
            ws = wb.create_sheet(title=title)
            for line in csv_text.strip().split('\n'):
                cells = []
                for c in line.split(','):
                    c = c.strip()
                    if c == '':
                        cells.append('')
                    else:
                        try:
                            cells.append(float(c) if '.' in c else int(c))
                        except ValueError:
                            cells.append(c)
                ws.append(cells)

        wb.save(path)
        return {'ok': True, 'path': path}

    def save_file_dialog(self, content, default_name):
        result = window.create_file_dialog(
            webview.SAVE_DIALOG,
            save_filename=default_name,
            file_types=('CSV Files (*.csv)',))
        if result:
            path = result if isinstance(result, str) else result[0]
            if not path.lower().endswith('.csv'):
                path += '.csv'
            with open(path, 'w', encoding='utf-8-sig', newline='') as f:
                f.write(content)
            return path
        return None

    # ─── 윈도우 제어 (frameless) ───
    _maximized = False

    def win_minimize(self):
        window.minimize()

    def win_toggle_max(self):
        if self._maximized:
            window.restore()
            self._maximized = False
        else:
            window.maximize()
            self._maximized = True

    def win_close(self):
        window.destroy()

    # ─── 유틸리티 ───
    def get_year_days(self, year):
        weekdays = ['월', '화', '수', '목', '금', '토', '일']
        days = []
        for m in range(1, 13):
            max_d = calendar.monthrange(year, m)[1]
            for d in range(1, max_d + 1):
                wd_idx = calendar.weekday(year, m, d)
                days.append({'m': m, 'd': d, 'wd': weekdays[wd_idx], 'wdi': wd_idx})
        return days


api = Api()
window = webview.create_window(
    '위드인천에너지 CHP 시뮬레이터',
    os.path.join(RESOURCE_DIR, 'index.html'),
    js_api=api, width=1440, height=920, min_size=(1000, 600),
    background_color='#0a0a0a',
    frameless=True, easy_drag=False
)


def apply_icon():
    """윈도우 생성 후 커스텀 아이콘 적용"""
    import time
    time.sleep(0.3)
    if sys.platform == 'win32':
        try:
            hwnd = ctypes.windll.user32.FindWindowW(None, '위드인천에너지 CHP 시뮬레이터')
            if hwnd:
                ico_path = os.path.join(RESOURCE_DIR, 'logo', 'with_logo.ico')
                hicon = ctypes.windll.user32.LoadImageW(
                    0, ico_path, 1, 0, 0, 0x00000010)
                if hicon:
                    ctypes.windll.user32.SendMessageW(hwnd, 0x0080, 0, hicon)
                    ctypes.windll.user32.SendMessageW(hwnd, 0x0080, 1, hicon)
        except Exception:
            pass


webview.start(debug=not getattr(sys, 'frozen', False), func=apply_icon)
